"""Export parks_latest.json to a formatted Excel workbook.

Usage (from project root):
    python -m data-pipeline.utils.export_excel                          # default output
    python -m data-pipeline.utils.export_excel -o my_parks.xlsx         # custom output path
    python -m data-pipeline.utils.export_excel -i data/final/other.json # custom input
"""

import argparse
import json
from pathlib import Path
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

_ROOT = Path(__file__).resolve().parent.parent.parent
DATA_DIR = _ROOT / "data" / "final"
DEFAULT_INPUT = DATA_DIR / "parks_latest.json"
DEFAULT_OUTPUT = _ROOT / "nc_parks.xlsx"

# Amenity columns derived from the canonical registry
_REGISTRY_PATH = _ROOT / "amenities.json"
_registry = json.loads(_REGISTRY_PATH.read_text(encoding="utf-8"))
AMENITY_COLS = [a["key"] for a in _registry]


def _pretty(name: str) -> str:
    """Convert snake_case amenity key to Title Case."""
    return name.replace("_", " ").title()


def export(input_path: Path, output_path: Path):
    parks = json.loads(input_path.read_text(encoding="utf-8"))

    # Sort by latitude then longitude so nearby parks are adjacent
    parks.sort(key=lambda p: (p.get("latitude") or 0, p.get("longitude") or 0))

    wb = Workbook()
    ws = wb.active
    ws.title = "Parks"

    # ---- Header row ----
    base_headers = [
        "Verified", "Satellite",
        "Name", "City", "County", "State", "Address",
        "Latitude", "Longitude", "Source", "Phone", "URL",
        "Google Maps", "Apple Maps",
    ]
    headers = base_headers + [_pretty(a) for a in AMENITY_COLS]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    thin_border = Border(
        bottom=Side(style="thin", color="CCCCCC"),
    )
    link_font = Font(color="2563EB", underline="single", size=10)
    check_font = Font(color="2E7D32", size=11)
    body_font = Font(size=10)
    body_align = Alignment(vertical="center", wrap_text=False)

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Data rows ----
    for row_idx, park in enumerate(parks, 2):
        lat = park.get("latitude")
        lon = park.get("longitude")
        name = park.get("name", "")
        url = park.get("url") or ""

        google_url = (f"https://www.google.com/maps/search/?api=1"
                      f"&query={lat},{lon}") if lat and lon else ""
        satellite_url = (f"https://www.google.com/maps/place/{quote_plus(name)}"
                         f"/@{lat},{lon},100m/data=!3m1!1e3") if lat and lon and name else ""
        apple_url = (f"https://maps.apple.com/?ll={lat},{lon}"
                     f"&q={name}") if lat and lon else ""

        base_values = [
            "",  # Verified — filled by dropdown
            satellite_url,
            name,
            park.get("city") or "",
            park.get("county") or "",
            park.get("state", "NC"),
            park.get("address") or "",
            lat,
            lon,
            park.get("source") or "",
            park.get("phone") or "",
            url,
            google_url,
            apple_url,
        ]

        amenities = park.get("amenities", {})
        amenity_values = ["✓" if amenities.get(a) else "" for a in AMENITY_COLS]

        for col_idx, val in enumerate(base_values + amenity_values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border

        # Format Satellite column as clickable hyperlink
        sat_col = base_headers.index("Satellite") + 1
        if satellite_url:
            cell = ws.cell(row=row_idx, column=sat_col)
            cell.hyperlink = satellite_url
            cell.value = "Open"
            cell.font = link_font

        # Format URL column as clickable hyperlink
        url_col = base_headers.index("URL") + 1
        if url:
            cell = ws.cell(row=row_idx, column=url_col)
            cell.hyperlink = url
            cell.font = link_font

        # Format Google Maps link
        gmap_col = base_headers.index("Google Maps") + 1
        if google_url:
            cell = ws.cell(row=row_idx, column=gmap_col)
            cell.hyperlink = google_url
            cell.value = "Open"
            cell.font = link_font

        # Format Apple Maps link
        amap_col = base_headers.index("Apple Maps") + 1
        if apple_url:
            cell = ws.cell(row=row_idx, column=amap_col)
            cell.hyperlink = apple_url
            cell.value = "Open"
            cell.font = link_font

        # Style the checkmarks
        for a_idx in range(len(AMENITY_COLS)):
            col = len(base_headers) + a_idx + 1
            cell = ws.cell(row=row_idx, column=col)
            if cell.value == "✓":
                cell.font = check_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---- Verified dropdown (Yes/No) ----
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    dv.prompt = "Select Yes or No"
    dv.promptTitle = "Verified"
    verified_col = get_column_letter(base_headers.index("Verified") + 1)
    dv.add(f"{verified_col}2:{verified_col}{len(parks) + 1}")
    ws.add_data_validation(dv)

    # ---- Column widths ----
    col_widths = {
        "Verified": 10, "Satellite": 10,
        "Name": 30, "City": 16, "County": 18, "State": 6,
        "Address": 35, "Latitude": 12, "Longitude": 12,
        "Source": 20, "Phone": 16, "URL": 30,
        "Google Maps": 12, "Apple Maps": 12,
    }
    for col_idx, h in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = col_widths.get(h, 10)

    # ---- Freeze panes & auto-filter ----
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(parks) + 1}"

    wb.save(output_path)
    print(f"Exported {len(parks)} parks to {output_path}")
    print(f"  Columns: {len(base_headers)} base + {len(AMENITY_COLS)} amenities")


def main():
    parser = argparse.ArgumentParser(description="Export parks to Excel")
    parser.add_argument("-i", "--input", type=Path, default=DEFAULT_INPUT,
                        help="Input JSON file")
    parser.add_argument("-o", "--output", type=Path, default=DEFAULT_OUTPUT,
                        help="Output Excel file")
    args = parser.parse_args()

    if not args.input.exists():
        print(f"Error: {args.input} not found. Run the pipeline first.")
        return

    export(args.input, args.output)


if __name__ == "__main__":
    main()
